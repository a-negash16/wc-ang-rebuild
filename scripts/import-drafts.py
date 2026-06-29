#!/usr/bin/env python3
import csv
import json
import os
import sys
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(sys.argv[1] if len(sys.argv) > 1 else "/Users/abrham/Downloads/drafts_v1.xlsx")
DRY_RUN = "--write" not in sys.argv

TEAM_ALIASES = {
    "england": "England",
    "usa": "USA",
    "united states": "USA",
    "cote divoire": "Côte d'Ivoire",
    "ivory coast": "Côte d'Ivoire",
    "dr congo": "Congo DR",
    "congo dr": "Congo DR",
    "congo": "Congo DR",
    "morroco": "Morocco",
    "bosnia": "Bosnia and Herzegovina",
    "cape verde": "Cabo Verde",
    "cabo verde": "Cabo Verde",
    "iran": "IR Iran",
}

PLAYER_ALIASES = {
    "mbappe": "Kylian Mbappe",
    "messi": "Lionel Messi",
    "ronaldo": "Cristiano Ronaldo",
    "yamal": "Lamine Yamal",
    "dembele": "Ousmane Dembele",
    "vini": "Vinicius Junior",
    "vinicius": "Vinicius Junior",
    "haaland": "Erling Haaland",
    "olise": "Michael Olise",
    "gakpo": "Cody Gakpo",
    "kane": "Harry Kane",
    "undav": "Deniz Undav",
    "diaz": "Luis Diaz",
    "bellingham": "Jude Bellingham",
    "bruno fernandes": "Bruno Fernandes",
    "de bruyne": "Kevin De Bruyne",
}

PLAYER_TEAM_ALIASES = {
    "balogun": "USA",
    "barcola": "France",
    "bellingham": "England",
    "bruno f": "Portugal",
    "bruno fernandes": "Portugal",
    "bruno g": "Brazil",
    "cunha": "Brazil",
    "de bruyne": "Belgium",
    "dembele": "France",
    "diaz": "Colombia",
    "doku": "Belgium",
    "gakpo": "Netherlands",
    "haaland": "Norway",
    "hakimi": "Morocco",
    "havertz": "Germany",
    "i sarr": "Senegal",
    "kane": "England",
    "l diaz": "Colombia",
    "manzambi": "Switzerland",
    "martinez": "Argentina",
    "mbappe": "France",
    "messi": "Argentina",
    "munoz": "Colombia",
    "olise": "France",
    "oyarzabal": "Spain",
    "pedri": "Spain",
    "pulisic": "USA",
    "quinonez": "Ecuador",
    "ronaldo": "Portugal",
    "saibari": "Morocco",
    "saibiri": "Morocco",
    "salah": "Egypt",
    "undav": "Germany",
    "vini": "Brazil",
    "vinicius": "Brazil",
    "yamal": "Spain",
}


def main():
    load_env(Path(".env.local"))
    require_env(["NEXT_PUBLIC_SUPABASE_URL", "SUPABASE_SERVICE_ROLE_KEY"])

    groups = rest_get("/groups?select=id,slug")
    managers = rest_get("/managers?select=id,display_name,groups(slug)&is_active=eq.true")
    teams = rest_get("/teams?select=id,name,fifa_code&limit=200")
    ensure_tables()

    group_by_slug = {row["slug"]: row for row in groups}
    manager_by_key = {
        (row.get("groups", {}).get("slug"), normalize(row["display_name"])): row
        for row in managers
    }
    team_by_name = {normalize(row["name"]): row for row in teams}
    for alias, canonical in TEAM_ALIASES.items():
        if normalize(canonical) in team_by_name:
            team_by_name[alias] = team_by_name[normalize(canonical)]

    team_rows, player_rows = read_workbook(WORKBOOK_PATH)
    diagnostics = []
    planned_team_rows = []
    planned_player_rows = []
    player_defs = {}

    for row in team_rows:
        group = group_by_slug.get(row["group_slug"])
        manager = manager_by_key.get((row["group_slug"], normalize(row["manager_name"])))
        team = team_by_name.get(normalize(row["drafted_team"]))
        if not group or not manager or not team:
            diagnostics.append({**row, "type": "team", "problem": missing_reason(group, manager, team)})
            continue
        planned_team_rows.append({
            "group_id": group["id"],
            "manager_id": manager["id"],
            "team_id": team["id"],
            "draft_slot": row["draft_slot"],
            "notes": None,
        })

    for row in player_rows:
        group = group_by_slug.get(row["group_slug"])
        manager = manager_by_key.get((row["group_slug"], normalize(row["manager_name"])))
        player_lookup_key = normalize(row["drafted_player"])
        player_name = PLAYER_ALIASES.get(player_lookup_key, row["drafted_player"].strip())
        player_team_name = PLAYER_TEAM_ALIASES.get(player_lookup_key)
        player_team = team_by_name.get(normalize(player_team_name)) if player_team_name else None
        if not group or not manager or not player_name:
            diagnostics.append({
                **row,
                "type": "player",
                "problem": missing_reason(group, manager, True, player_name),
            })
            continue
        player_key = player_name
        player_defs[player_key] = {
            "team_id": player_team["id"] if player_team else None,
            "display_name": player_name,
            "position": None,
            "external_player_id": None,
        }
        planned_player_rows.append({
            "group_id": group["id"],
            "manager_id": manager["id"],
            "player_key": player_key,
            "draft_slot": row["draft_slot"],
            "notes": None,
        })

    if not DRY_RUN:
        for player in player_defs.values():
            rest_post("/players?on_conflict=display_name", [player], prefer="resolution=merge-duplicates")
        players = rest_get("/players?select=id,display_name&limit=500")
        player_id_by_key = {row["display_name"]: row["id"] for row in players}
        drafted_players = []
        for row in planned_player_rows:
            drafted_players.append({
                "group_id": row["group_id"],
                "manager_id": row["manager_id"],
                "player_id": player_id_by_key[row["player_key"]],
                "draft_slot": row["draft_slot"],
                "notes": row["notes"],
            })

        rest_post("/drafted_teams?on_conflict=group_id,manager_id,team_id", planned_team_rows, prefer="resolution=merge-duplicates")
        rest_post("/drafted_players?on_conflict=group_id,manager_id,player_id", drafted_players, prefer="resolution=merge-duplicates")
        tally_rows = [{"player_id": player_id} for player_id in sorted({row["player_id"] for row in drafted_players})]
        rest_post("/player_stat_tallies?on_conflict=player_id", tally_rows, prefer="resolution=merge-duplicates")

    print(json.dumps({
        "mode": "dry-run" if DRY_RUN else "write",
        "workbook": str(WORKBOOK_PATH),
        "drafted_teams": len(planned_team_rows),
        "drafted_players": len(planned_player_rows),
        "players": len(player_defs),
        "unmatched": diagnostics,
    }, indent=2))


def read_workbook(path):
    wb = load_workbook(path, data_only=True)
    team_rows = []
    player_rows = []
    for ws in wb.worksheets:
        headers = [cell.value for cell in ws[1]]
        if "group_slug" not in headers or "manager_name" not in headers:
            continue
        rows = team_rows if "drafted_team" in headers else player_rows if "drafted_player" in headers else None
        if rows is None:
            continue
        for values in ws.iter_rows(min_row=2, values_only=True):
            item = dict(zip(headers, values))
            if not item.get("group_slug") or not item.get("manager_name"):
                continue
            draft_value = item.get("drafted_team") or item.get("drafted_player")
            if not draft_value:
                continue
            if normalize(draft_value) in {"missing", "tbd", "to be added"}:
                continue
            rows.append({
                "group_slug": str(item["group_slug"]).strip(),
                "manager_name": str(item["manager_name"]).strip(),
                "drafted_team": str(item.get("drafted_team") or "").strip(),
                "drafted_player": str(item.get("drafted_player") or "").strip(),
                "draft_slot": int(item["draft_slot"]) if item.get("draft_slot") not in (None, "") else None,
                "sheet": ws.title,
            })
    return team_rows, player_rows


def ensure_tables():
    for table in ["players", "drafted_teams", "drafted_players", "player_stat_tallies"]:
        try:
            rest_get(f"/{table}?select=id&limit=1")
        except RuntimeError as exc:
            raise SystemExit(
                f"Draft table check failed for {table}. Apply supabase/generated/apply-all.sql "
                f"or migration 0003_draft_scoring.sql first. Details: {exc}"
            )


def normalize(value):
    value = str(value or "").strip().lower()
    return " ".join(value.replace("’", "'").replace(".", "").split())


def missing_reason(group, manager, team, player_name=True):
    missing = []
    if not group:
        missing.append("group")
    if not manager:
        missing.append("manager")
    if not team:
        missing.append("team")
    if not player_name:
        missing.append("player_alias")
    return "missing_" + "_".join(missing)


def load_env(path):
    if not path.exists():
        return
    for line in path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip("\"'"))


def require_env(names):
    missing = [name for name in names if not os.environ.get(name)]
    if missing:
        raise SystemExit(f"Missing env vars: {', '.join(missing)}")


def rest_url(path):
    return os.environ["NEXT_PUBLIC_SUPABASE_URL"].rstrip("/") + "/rest/v1" + path


def rest_headers(prefer=None):
    key = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
    }
    if prefer:
        headers["Prefer"] = prefer
    return headers


def rest_get(path):
    request = urllib.request.Request(rest_url(path), headers=rest_headers())
    try:
        with urllib.request.urlopen(request) as response:
            return json.loads(response.read().decode() or "[]")
    except urllib.error.HTTPError as exc:
        raise RuntimeError(exc.read().decode())


def rest_post(path, rows, prefer="resolution=merge-duplicates"):
    if not rows:
        return
    request = urllib.request.Request(
        rest_url(path),
        data=json.dumps(rows).encode(),
        headers=rest_headers(prefer=f"{prefer},return=minimal"),
        method="POST",
    )
    try:
        with urllib.request.urlopen(request) as response:
            response.read()
    except urllib.error.HTTPError as exc:
        raise RuntimeError(exc.read().decode())


if __name__ == "__main__":
    main()
